"""
Adiciona aba 'Precificação' ao relatorio_jira_completo.xlsx
Preços reais extraídos do painel Atlassian (Março 2026)
"""
import pandas as pd
import warnings
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

# ── Preços reais Atlassian (print do painel) ──────────────────────────────────
# Valor mensal total / quantidade de licenças ativas = custo por usuário/mês
PRECOS = {
    # produto                    : (total_USD,  qtd_licencas)
    'Jira'                       : (1907.20,    220),
    'Confluence'                 : (1252.50,    205),
    'Jira Service Management'    : (1258.43,     54),
    'Jira Product Discovery'     : ( 711.18,     67),
    'Compass'                    : (  38.19,      4),
}

# Custo por usuário/mês (USD)
CUSTO_USER = {p: round(total / qtd, 4) for p, (total, qtd) in PRECOS.items()}
print('Custo por usuário/mês (USD):')
for p, c in CUSTO_USER.items():
    print(f'  {p}: US$ {c:.2f}')

# ── Carregar dados ────────────────────────────────────────────────────────────
df = pd.read_excel('relatorio_jira_completo.xlsx', sheet_name='Relatorio Completo')
df['Diretoria'] = df['Diretoria'].fillna('NÃO LOCALIZADO').str.upper().str.strip()

def has_user(val):
    return pd.notna(val) and 'User' in str(val)

# ── Custo por usuário ─────────────────────────────────────────────────────────
for p, custo in CUSTO_USER.items():
    col = f'_custo_{p}'
    df[col] = df[p].apply(lambda x: custo if has_user(x) else 0.0)

df['Custo Total Mensal (USD)'] = sum(df[f'_custo_{p}'] for p in CUSTO_USER)

# ── Resumo por Diretoria ──────────────────────────────────────────────────────
agg = {}
for d, grp in df.groupby('Diretoria'):
    row = {'Diretoria': d, 'Usuários': len(grp)}
    total = 0.0
    for p, custo in CUSTO_USER.items():
        n = int(grp[p].apply(has_user).sum())
        val = round(n * custo, 2)
        row[f'Usuários {p}'] = n
        row[f'Custo {p} (USD)'] = val
        total += val
    row['TOTAL MENSAL (USD)'] = round(total, 2)
    agg[d] = row

df_prec = pd.DataFrame(list(agg.values()))
df_prec = df_prec.sort_values('TOTAL MENSAL (USD)', ascending=False).reset_index(drop=True)

# Linha de totais
totais = {'Diretoria': 'TOTAL GERAL', 'Usuários': len(df)}
for p, (total_real, _) in PRECOS.items():
    totais[f'Usuários {p}'] = int(df[p].apply(has_user).sum())
    totais[f'Custo {p} (USD)'] = total_real
totais['TOTAL MENSAL (USD)'] = sum(PRECOS[p][0] for p in PRECOS)
df_prec = pd.concat([df_prec, pd.DataFrame([totais])], ignore_index=True)

# ── Resumo executivo simples (uma linha por produto) ─────────────────────────
resumo_rows = []
for p, (total_real, qtd) in PRECOS.items():
    resumo_rows.append({
        'Produto'               : p,
        'Plano'                 : 'Standard',
        'Usuários Licenciados'  : qtd,
        'Custo Mensal (USD)'    : total_real,
        'Custo/Usuário (USD)'   : round(total_real / qtd, 2),
        'Custo Anual (USD)'     : round(total_real * 12, 2),
    })
df_resumo = pd.DataFrame(resumo_rows)
total_mensal = sum(r['Custo Mensal (USD)'] for r in resumo_rows)
df_resumo = pd.concat([df_resumo, pd.DataFrame([{
    'Produto': 'TOTAL',
    'Plano': '',
    'Usuários Licenciados': '',
    'Custo Mensal (USD)': total_mensal,
    'Custo/Usuário (USD)': '',
    'Custo Anual (USD)': round(total_mensal * 12, 2),
}])], ignore_index=True)

# ── Escrever no Excel existente ───────────────────────────────────────────────
FILE = 'relatorio_jira_completo.xlsx'

# Carrega workbook existente e remove abas antigas se já existirem
wb = load_workbook(FILE)
for name in ['Precificação', 'Resumo Custos']:
    if name in wb.sheetnames:
        del wb[name]

# Estilos
H_FILL  = PatternFill('solid', fgColor='1F3864')
H_FONT  = Font(bold=True, color='FFFFFF', size=10)
TOT_FILL= PatternFill('solid', fgColor='2E4057')
TOT_FONT= Font(bold=True, color='DEFF9A', size=10)
ALT1    = PatternFill('solid', fgColor='FFFFFF')
ALT2    = PatternFill('solid', fgColor='EEF3FB')
MONEY   = PatternFill('solid', fgColor='E8F5E9')
BORDER  = Border(*[Side(style='thin', color='BFBFBF')]*4)
USD_FMT = '"US$ "#,##0.00'

def auto_col_width(ws):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 3, 40)

def write_df(wb, sheet_name, df_in, money_cols=None, total_last=True):
    ws = wb.create_sheet(title=sheet_name)
    cols = list(df_in.columns)

    # Header
    for ci, col in enumerate(cols, 1):
        c = ws.cell(1, ci, col)
        c.fill = H_FILL; c.font = H_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 36

    n_rows = len(df_in)
    for ri, (_, row) in enumerate(df_in.iterrows(), 2):
        is_total = total_last and ri == n_rows + 1
        fill = TOT_FILL if is_total else (ALT2 if ri % 2 == 0 else ALT1)
        for ci, col in enumerate(cols, 1):
            val = row[col]
            c   = ws.cell(ri, ci, value=val if not pd.isna(val) else None)
            c.fill   = fill
            c.font   = TOT_FONT if is_total else Font(size=10)
            c.border = BORDER
            c.alignment = Alignment(vertical='center',
                                    horizontal='right' if isinstance(val, (int, float)) else 'left')
            if money_cols and col in money_cols and isinstance(val, (int, float)):
                c.number_format = USD_FMT
                if not is_total:
                    c.fill = PatternFill('solid', fgColor='F0F7F0')

    auto_col_width(ws)
    ws.freeze_panes = 'B2'
    return ws

# ── Aba: Resumo Custos ────────────────────────────────────────────────────────
money_res = ['Custo Mensal (USD)', 'Custo/Usuário (USD)', 'Custo Anual (USD)']
write_df(wb, 'Resumo Custos', df_resumo, money_cols=money_res)

# ── Aba: Precificação por Diretoria ───────────────────────────────────────────
money_prec = [f'Custo {p} (USD)' for p in CUSTO_USER] + ['TOTAL MENSAL (USD)']
write_df(wb, 'Precificação', df_prec, money_cols=money_prec)

wb.save(FILE)
print(f'\nAbas adicionadas: "Resumo Custos" e "Precificação"')
print(f'Total mensal: US$ {total_mensal:,.2f} | Anual: US$ {total_mensal*12:,.2f}')
print('\nCusto por diretoria (top 5):')
for _, r in df_prec.head(5).iterrows():
    print(f'  {r["Diretoria"]}: US$ {r["TOTAL MENSAL (USD)"]:,.2f}/mês ({int(r["Usuários"])} usuários)')
