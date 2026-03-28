"""
relatorio_jira_completo.xlsx
  Base  : JIRA_separado.xlsx (Resultado) — e-mail + Diretoria já mapeados
  RH    : rh_atual.xlsx → verifica e corrige Centro de Custo; traz Cargo/Area/Gestor
           se não encontrado → mantém CCusto do JIRA_separado + marca "CONFERIR"
  Jira  : export.csv → todas as licenças do usuário (por e-mail)
  Sem OBS
"""
import pandas as pd
import unicodedata
import warnings
warnings.filterwarnings('ignore')
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── Helpers ───────────────────────────────────────────────────────────────────
def norm(s):
    if pd.isna(s): return ''
    s = str(s).upper()
    s = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in s if not unicodedata.combining(c)).strip()

def first_last(name):
    p = norm(name).split()
    return (p[0]+' '+p[-1]) if len(p) >= 2 else norm(name)

def first2_last(name):
    p = norm(name).split()
    return (p[0]+' '+p[1]+' '+p[-1]) if len(p) >= 3 else norm(name)

# ── Carregar dados ────────────────────────────────────────────────────────────
df_jira = pd.read_csv('export.csv', encoding='utf-8', sep=None, engine='python')
df_rh   = pd.read_excel('rh_atual.xlsx')
df_res  = pd.read_excel('JIRA_separado.xlsx', sheet_name='Resultado')


# ── Mapa de e-mail → Diretoria/CCusto do JIRA_separado (para lookup) ─────────
df_res['_email_key'] = df_res['E-mail'].str.lower().str.strip()
email_to_dir = df_res.set_index('_email_key')[['Diretoria', 'Centro de Custo', 'Nome', 'Status da Licença']].to_dict('index')

# ── Normalizar chaves de nome ─────────────────────────────────────────────────
for df, col in [(df_rh, 'NOME'), (df_res, 'Nome')]:
    df['_norm'] = df[col].apply(norm)
    df['_fl']   = df[col].apply(first_last)
    df['_f2l']  = df[col].apply(first2_last)

RH_COLS = ['CCUSTO', 'DIRETORIA', 'CARGO (SEM SENIORIDADE)', 'ÁREA', 'GESTOR']

def rh_lookup(key_val, key_col):
    rows = df_rh[df_rh[key_col] == key_val]
    return rows.iloc[0][RH_COLS].to_dict() if len(rows) else {}

# ── Base: export.csv como fonte principal (todos os 228 usuários) ─────────────
PROD_COLS = {
    'Jira - infoglobo'                    : 'Jira',
    'Confluence - infoglobo'              : 'Confluence',
    'Jira Service Management - infoglobo' : 'Jira Service Management',
    'Jira Product Discovery - infoglobo'  : 'Jira Product Discovery',
    'Assets - infoglobo'                  : 'Assets',
    'Compass - infoglobo'                 : 'Compass',
    'Atlas - infoglobo'                   : 'Atlas',
    'Goals - infoglobo'                   : 'Goals',
    'Projects - infoglobo'                : 'Projects',
}

j = df_jira[['User name', 'email', 'User status',
             'Last seen in Jira - infoglobo',
             'Last seen in Confluence - infoglobo',
             'Last seen in Jira Service Management - infoglobo'] +
            list(PROD_COLS.keys())].copy()
j.rename(columns={'User name': 'Nome',
                  'email': 'E-mail',
                  'User status': 'Status da Licença',
                  'Last seen in Jira - infoglobo': 'Último acesso Jira',
                  'Last seen in Confluence - infoglobo': 'Último acesso Confluence',
                  'Last seen in Jira Service Management - infoglobo': 'Último acesso JSM',
                  **PROD_COLS}, inplace=True)
j['_email'] = j['E-mail'].str.lower().str.strip()

# Enriquecer com Diretoria/CCusto do JIRA_separado via e-mail
j['Diretoria']       = j['_email'].map(lambda e: email_to_dir.get(e, {}).get('Diretoria', ''))
j['Centro de Custo'] = j['_email'].map(lambda e: email_to_dir.get(e, {}).get('Centro de Custo', ''))

# Cruzar com RH por nome (usando Nome do export)
j['_norm'] = j['Nome'].apply(norm)
j['_fl']   = j['Nome'].apply(first_last)
j['_f2l']  = j['Nome'].apply(first2_last)

rh_data = []
for _, row in j.iterrows():
    hit = (rh_lookup(row['_norm'], '_norm')
           or rh_lookup(row['_fl'],  '_fl')
           or rh_lookup(row['_f2l'], '_f2l'))
    rh_data.append(hit)

rh_df = pd.DataFrame(rh_data, index=j.index)
base  = pd.concat([j.reset_index(drop=True), rh_df.reset_index(drop=True)], axis=1)

# ── Validação cruzada: nome + CCusto ─────────────────────────────────────────
def normalizar_cc(v):
    if pd.isna(v): return ''
    try:
        return str(int(float(str(v).strip())))  # 1142225000.0 → "1142225000"
    except (ValueError, TypeError):
        return str(v).strip().upper()

def status_conferir(row):
    rh_ccusto   = normalizar_cc(row.get('CCUSTO'))
    jira_ccusto = normalizar_cc(row.get('Centro de Custo'))
    if not rh_ccusto:
        return 'CONFERIR'                               # nome não encontrado no RH
    if not jira_ccusto:
        return ''                                       # sem CCusto no JIRA — aceita o do RH
    if rh_ccusto == jira_ccusto:
        return ''                                       # OK — batem
    return f'DIVERGÊNCIA (RH:{rh_ccusto} / JIRA:{jira_ccusto})'  # CCusto diferente

base['Conferir CCusto'] = base.apply(status_conferir, axis=1)

final = base.copy()

# ── Montar saída final ────────────────────────────────────────────────────────
# Relatorio Completo: sem coluna Conferir, CCusto vem do JIRA_separado
out = final[[
    'Nome',
    'E-mail',
    'Status da Licença',
    'Diretoria',
    'Centro de Custo',
    'CARGO (SEM SENIORIDADE)',
    'ÁREA',
    'GESTOR',
    'Jira',
    'Confluence',
    'Jira Service Management',
    'Jira Product Discovery',
    'Assets',
    'Compass',
    'Atlas',
    'Goals',
    'Projects',
    'Último acesso Jira',
    'Último acesso Confluence',
    'Último acesso JSM',
]].copy()

out.rename(columns={
    'Status da Licença'       : 'Status',
    'CARGO (SEM SENIORIDADE)' : 'Cargo',
    'ÁREA'                    : 'Area',
    'GESTOR'                  : 'Gestor',
}, inplace=True)

# ── Estilo Excel ──────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill('solid', fgColor='1F3864')
HDR_FONT  = Font(bold=True, color='FFFFFF', size=10)
EVEN_FILL = PatternFill('solid', fgColor='DCE6F1')
ODD_FILL  = PatternFill('solid', fgColor='FFFFFF')
WARN_FILL = PatternFill('solid', fgColor='FFE699')   # amarelo — CONFERIR
DIV_FILL  = PatternFill('solid', fgColor='FFD0D0')   # vermelho claro — DIVERGÊNCIA
BORDER    = Border(
    left   = Side(style='thin', color='BFBFBF'),
    right  = Side(style='thin', color='BFBFBF'),
    top    = Side(style='thin', color='BFBFBF'),
    bottom = Side(style='thin', color='BFBFBF'),
)

def style_sheet(ws, warn_col=None, col_list=None):
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 32

    warn_idx = None
    if warn_col and col_list:
        try:
            warn_idx = col_list.index(warn_col) + 1
        except ValueError:
            pass

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
        fill = EVEN_FILL if r_idx % 2 == 0 else ODD_FILL
        status_val = row[warn_idx-1].value if warn_idx else ''
        if status_val == 'CONFERIR':
            row_fill = WARN_FILL   # amarelo
        elif status_val and status_val.startswith('DIVERGÊNCIA'):
            row_fill = DIV_FILL    # vermelho claro
        else:
            row_fill = fill
        for cell in row:
            cell.fill = row_fill
            cell.alignment = Alignment(vertical='center')
            cell.border = BORDER

    for col in ws.columns:
        max_w = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max(min(max_w + 2, 45), 10)
    ws.freeze_panes = 'A2'

OUTPUT = 'relatorio_jira_completo.xlsx'

# Aba Conferir: usuários não encontrados no RH, com coluna de marcação
conferir_base = final[final['Conferir CCusto'] != ''].copy()
conferir = conferir_base[[
    'Nome', 'E-mail', 'Status da Licença', 'Diretoria', 'Centro de Custo', 'Conferir CCusto',
    'Jira', 'Confluence', 'Jira Service Management', 'Jira Product Discovery',
    'Assets', 'Compass', 'Atlas', 'Goals', 'Projects',
    'Último acesso Jira', 'Último acesso Confluence', 'Último acesso JSM',
]].copy()
conferir.rename(columns={'Status da Licença': 'Status'}, inplace=True)

with pd.ExcelWriter(OUTPUT, engine='openpyxl') as writer:
    # Aba principal: todos os usuários, sem marcação
    out.to_excel(writer, sheet_name='Relatorio Completo', index=False)
    # Aba de conferência: só os pendentes, para correção manual
    conferir.to_excel(writer, sheet_name='Conferir CCusto', index=False)

    style_sheet(writer.sheets['Relatorio Completo'],
                col_list=list(out.columns))
    style_sheet(writer.sheets['Conferir CCusto'],
                warn_col='Conferir CCusto', col_list=list(conferir.columns))

print(f'Arquivo  : {OUTPUT}')
print(f'Relatorio Completo: {len(out)} usuarios')
print(f'Conferir CCusto   : {len(conferir)} usuarios')
print(f'Licenças por produto:')
for p in ['Jira','Confluence','Jira Service Management','Jira Product Discovery','Assets']:
    n = out[p].apply(lambda x: pd.notna(x) and str(x).strip() not in ('',)).sum()
    print(f'  {p}: {n}')
